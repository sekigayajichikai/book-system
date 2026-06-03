"""
Excel → Google Sheets 取込スクリプト

Excelのマトリクスデータを読み取り、Google Sheets APIで
スプレッドシートBに直接書き込む。

使い方:
  python excel_to_sheets.py                        # 当月＋翌月を取込
  python excel_to_sheets.py --year 2026 --month 5  # 指定月を取込
  python excel_to_sheets.py --all                  # 全シートを取込
  python excel_to_sheets.py --dry-run              # 実際には書き込まず確認のみ
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import argparse
import json
import time
import openpyxl
from datetime import datetime
from pathlib import Path

# === 設定 ===
SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / '会館日程表（新）.xlsx'
SPREADSHEET_ID = '1_FI98IB9ZZHmzjQa_LA94glLWH9uHOj0Fe30eGcTE7s'  # スプレッドシートB（Excel取込用）

# イベント名→主催団体の簡易マッピング
ORG_MAP = {
    '囲碁': '自主活動部',
    'カラオケ': '自主活動部',
    '関ヶ谷クラブ': '自主活動部',
    'ディスクコンサート': '自主活動部',
    '図書': '自主活動部',
    'ふれあい': '自主活動部',
    'ブルーベル': '自主活動部',
    'ブル―ベル': '自主活動部',
    'トーンチャイム': '自主活動部',
    'ちりとてちん': '自主活動部',
    'オペラ': '自主活動部',
    '読書': '自主活動部',
    'つなぎの会': '自主活動部',
    'ききょう': '自主活動部',
    '見まわり隊': '自主活動部',
    '役員': '役員',
    '総会': '役員',
    '新役員': '役員',
    '事務局': '事務局',
    '会館予約': '事務局',
    '会計監査': '事務局',
    '防災': '委員会',
    'HP': '委員会',
    'DX': '委員会',
    '規約': '委員会',
    '広報': '委員会',
    '建築協定': '委員会',
    '環境': '委員会',
    '地区長': '地区長・班長',
    '班長': '地区長・班長',
    '合同会議': '地区長・班長',
}


def guess_org(title):
    """イベント名から主催団体を推測"""
    for keyword, org in ORG_MAP.items():
        if keyword in title:
            return org
    return ''


# === Excel読み取り ===

def find_month_sheets(wb):
    """ワークブックから月別シートを検出"""
    results = []
    for name in wb.sheetnames:
        ws = wb[name]
        year_val = ws.cell(row=1, column=9).value
        month_val = ws.cell(row=1, column=12).value
        if year_val and month_val:
            try:
                y = int(year_val)
                m = int(month_val)
                if 2020 <= y <= 2099 and 1 <= m <= 12:
                    results.append((y, m, name))
            except (ValueError, TypeError):
                continue
    return results


def read_month_events(wb, sheet_name):
    """シートからイベント一覧を抽出"""
    ws = wb[sheet_name]
    year = int(ws.cell(row=1, column=9).value)
    month = int(ws.cell(row=1, column=12).value)

    # 日付列を取得
    date_cols = []
    for col in range(4, ws.max_column + 1):
        v = ws.cell(row=2, column=col).value
        if isinstance(v, datetime) and v.year >= 2000 and v.month == month:
            date_cols.append((col, v))

    # 行定義
    rows_def = [
        (4, '午前', '会議室'),
        (5, '午前', '和室（畳側）'),
        (6, '午前', '和室（椅子側）'),
        (7, '午前', '図書室'),
        (9, '午後', '会議室'),
        (10, '午後', '和室（畳側）'),
        (11, '午後', '和室（椅子側）'),
        (12, '午後', '図書室'),
        (13, '夜間', ''),
    ]

    events = []
    for row_num, slot, room in rows_def:
        for col, dt in date_cols:
            v = ws.cell(row=row_num, column=col).value
            if not v:
                continue
            title = str(v).strip().replace('\u3000', '')
            if not title or title == '×':
                continue

            events.append({
                'year': year,
                'month': month,
                'day': dt.day,
                'slot': slot,
                'room': room,
                'title': title,
                'org': guess_org(title),
            })

    return events


# === Google Sheets API ===

# 時間帯定義
TIME_SLOTS = {
    '午前': ('09:00', '12:00'),
    '午後': ('13:00', '16:00'),
    '夜間': ('17:00', '20:00'),
}

_sheets_service = None

def get_sheets_service():
    """Google Sheets APIサービスを取得"""
    global _sheets_service
    if _sheets_service:
        return _sheets_service

    import os
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    TOKEN_PATH = SCRIPT_DIR / '.sheets_token.json'
    CREDS_PATH = os.path.expandvars(
        r'%APPDATA%\gcloud\application_default_credentials.json'
    )

    creds = None
    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            # gcloud ADCのclient_id/secretを流用してOAuthフロー
            import json
            with open(CREDS_PATH) as f:
                adc = json.load(f)
            flow = InstalledAppFlow.from_client_config(
                {
                    'installed': {
                        'client_id': adc['client_id'],
                        'client_secret': adc['client_secret'],
                        'auth_uri': 'https://accounts.google.com/o/oauth2/auth',
                        'token_uri': 'https://oauth2.googleapis.com/token',
                    }
                },
                SCOPES,
            )
            creds = flow.run_local_server(port=0)
        with open(TOKEN_PATH, 'w') as f:
            f.write(creds.to_json())

    _sheets_service = build('sheets', 'v4', credentials=creds)
    return _sheets_service


def get_sheet_name():
    """スプレッドシートBの最初のシート名を取得"""
    service = get_sheets_service()
    result = service.spreadsheets().get(
        spreadsheetId=SPREADSHEET_ID,
        fields='sheets.properties.title'
    ).execute()
    return result['sheets'][0]['properties']['title']


def load_existing(year, month):
    """スプレッドシートBの既存データを取得"""
    service = get_sheets_service()
    sheet_name = get_sheet_name()

    try:
        result = service.spreadsheets().values().get(
            spreadsheetId=SPREADSHEET_ID,
            range=f'{sheet_name}!A:N',
        ).execute()
        rows = result.get('values', [])
    except Exception as e:
        print(f'  既存データ取得失敗: {e}')
        return {}

    if len(rows) <= 1:
        return {}

    headers = rows[0]
    existing = {}
    for row in rows[1:]:
        if len(row) < 10:
            continue
        # パディング
        while len(row) < len(headers):
            row.append('')

        date_str = row[1]  # 日付列
        slot = row[8]      # 区分列
        room = row[7]      # 部屋列
        title = row[9]     # イベント名列
        status = row[12] if len(row) > 12 else ''

        if status == 'キャンセル':
            continue

        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            if dt.year == year and dt.month == month:
                key = f'{slot}_{room}_{dt.day}'
                existing[key] = {'title': title}
        except (ValueError, TypeError):
            continue

    return existing


def save_events_batch(events):
    """複数イベントを一括でスプレッドシートBに追加"""
    if not events:
        return 0

    service = get_sheets_service()
    sheet_name = get_sheet_name()

    rows = []
    for evt in events:
        date_str = f'{evt["year"]}-{evt["month"]:02d}-{evt["day"]:02d}'
        start_time, end_time = TIME_SLOTS.get(evt['slot'], ('', ''))
        event_id = f'excel_{date_str.replace("-","")}_{evt["title"][:10]}'
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        rows.append([
            event_id,           # id
            date_str,           # 日付
            'FALSE',            # 終日
            start_time,         # 開始時刻
            end_time,           # 終了時刻
            evt['org'],         # 主催団体
            '自治会館',          # 会場
            evt['room'],        # 部屋
            evt['slot'],        # 区分
            evt['title'],       # イベント名
            '',                 # 備考
            now_str,            # 更新日時
            '確定',             # ステータス
            'Excel取込',        # 事務局メモ
        ])

    result = service.spreadsheets().values().append(
        spreadsheetId=SPREADSHEET_ID,
        range=f'{sheet_name}!A:N',
        valueInputOption='USER_ENTERED',
        insertDataOption='INSERT_ROWS',
        body={'values': rows},
    ).execute()

    updated = result.get('updates', {}).get('updatedRows', 0)
    return updated


# === メイン処理 ===

def sync_month(wb, sheet_name, year, month, dry_run=False):
    """1ヶ月分のExcelデータをSheetsに同期"""
    print(f'\n=== {year}年{month}月 ({sheet_name.strip()}) ===')

    # Excelからイベント読取
    excel_events = read_month_events(wb, sheet_name)
    print(f'  Excel: {len(excel_events)}件')

    # 既存データ取得
    existing = load_existing(year, month)
    print(f'  確定シート: {len(existing)}件')

    # 差分を計算
    new_events = []
    skip_events = []
    for evt in excel_events:
        key = f'{evt["slot"]}_{evt["room"]}_{evt["day"]}'
        if key in existing:
            ex = existing[key]
            if ex['title'] == evt['title']:
                skip_events.append(evt)
            else:
                # タイトルが違う → 更新
                new_events.append(evt)
        else:
            new_events.append(evt)

    print(f'  スキップ（既存と同じ）: {len(skip_events)}件')
    print(f'  新規/更新: {len(new_events)}件')

    if dry_run:
        for evt in new_events:
            print(f'    [DRY] {evt["month"]}/{evt["day"]} {evt["slot"]} {evt["room"]} → {evt["title"]} ({evt["org"]})')
        return len(new_events)

    # 一括保存
    try:
        count = save_events_batch(new_events)
        print(f'  結果: {count}行 書き込み成功')
        return count
    except Exception as e:
        print(f'  書き込み失敗: {e}')
        return 0


def main():
    parser = argparse.ArgumentParser(description='Excel → Google Sheets 取込')
    parser.add_argument('--year', type=int, help='対象年')
    parser.add_argument('--month', type=int, help='対象月')
    parser.add_argument('--all', action='store_true', help='全シートを取込')
    parser.add_argument('--dry-run', action='store_true', help='確認のみ（書き込まない）')
    args = parser.parse_args()

    print(f'Excel: {XLSX_PATH}')
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    sheets = find_month_sheets(wb)
    print(f'シート数: {len(sheets)}')

    if args.all:
        targets = sheets
    elif args.year and args.month:
        targets = [(y, m, n) for y, m, n in sheets if y == args.year and m == args.month]
        if not targets:
            print(f'{args.year}年{args.month}月のシートが見つかりません')
            return
    else:
        # デフォルト: 当月＋翌月
        now = datetime.now()
        cur = (now.year, now.month)
        nxt = (now.year + 1, 1) if now.month == 12 else (now.year, now.month + 1)
        targets = [(y, m, n) for y, m, n in sheets if (y, m) in [cur, nxt]]

    total = 0
    for year, month, sheet_name in targets:
        total += sync_month(wb, sheet_name, year, month, dry_run=args.dry_run)

    print(f'\n完了: 合計 {total}件 {"（dry-run）" if args.dry_run else ""}')


if __name__ == '__main__':
    main()
