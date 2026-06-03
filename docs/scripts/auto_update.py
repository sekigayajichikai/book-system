"""
Excel → HTML → GitHub Pages 自動更新スクリプト

Windowsタスクスケジューラで10分おきに実行する想定。
Excelファイルの更新日時を監視し、変更があればHTMLを再生成してGitHubにpush。

使い方:
  pythonw auto_update.py          （バックグラウンド実行、コンソール非表示）
  python  auto_update.py --force  （強制実行、デバッグ用）

初回セットアップ:
  1. config.json を作成（config.example.json を参考に）
  2. タスクスケジューラに登録
"""
import sys
import os
import json
import base64
import hashlib
import logging
from pathlib import Path
from datetime import datetime

# UTF-8出力（pythonw.exeでも安全に動くように）
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

import openpyxl
import urllib.request
import urllib.error

# === パス設定 ===
SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / '会館日程表（新）.xlsx'
CONFIG_PATH = SCRIPT_DIR / 'config.json'
STATE_PATH = SCRIPT_DIR / '.last_update_state.json'
LOG_PATH = SCRIPT_DIR / 'auto_update.log'

# === ログ設定 ===
logging.basicConfig(
    filename=str(LOG_PATH),
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    encoding='utf-8',
)
log = logging.getLogger(__name__)


def load_config():
    """config.json を読み込む"""
    if not CONFIG_PATH.exists():
        log.error('config.json が見つかりません')
        sys.exit(1)
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_file_hash(path):
    """ファイルのMD5ハッシュを返す"""
    h = hashlib.md5()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            h.update(chunk)
    return h.hexdigest()


def has_changed():
    """Excelファイルが前回から変更されたか判定"""
    if not XLSX_PATH.exists():
        log.error(f'Excelファイルが見つかりません: {XLSX_PATH}')
        return False

    current_hash = get_file_hash(XLSX_PATH)

    if STATE_PATH.exists():
        with open(STATE_PATH, 'r') as f:
            state = json.load(f)
        if state.get('hash') == current_hash:
            return False

    return True


def save_state():
    """更新状態を保存"""
    state = {
        'hash': get_file_hash(XLSX_PATH),
        'updated_at': datetime.now().isoformat(),
    }
    with open(STATE_PATH, 'w') as f:
        json.dump(state, f)


# =============================================
# Excel → HTML 生成（generate_excel_style_html.py から流用）
# =============================================

COLORS = {
    '会議室':  '#FFFF99',
    '和室':    '#E0EFF5',
    '洋室':    '#E0EFF5',
    '図書室':  '#EE9CD7',
    '夜間':    '#FFFFFF',
}
LABEL_COLORS = {
    '会議室':  '#FFFF99',
    '和室':    '#B6DDE8',
    '洋室':    '#B6DDE8',
    '図書室':  '#EE9CD7',
}
DOW_NAMES = ['月', '火', '水', '木', '金', '土', '日']
HALF = 15

# シート名→(年, 月) のマッピングを動的に作る
def find_month_sheets(wb):
    """ワークブックから月別シートを検出して (year, month, sheet_name) のリストを返す"""
    import re
    results = []
    for name in wb.sheetnames:
        ws = wb[name]
        # row1に年月がある
        year_val = ws.cell(row=1, column=9).value
        month_val = ws.cell(row=1, column=12).value
        if year_val and month_val:
            try:
                y = int(year_val)
                m = int(month_val)
                if 2020 <= y <= 2099 and 1 <= m <= 12:
                    results.append((y, m, name))
            except (ValueError, TypeError):
                continue
    return results


def read_month_data(wb, sheet_name):
    ws = wb[sheet_name]
    year = int(ws.cell(row=1, column=9).value)
    month = int(ws.cell(row=1, column=12).value)
    memo = ws.cell(row=1, column=15).value or ''

    date_cols = []
    for col in range(4, ws.max_column + 1):
        v = ws.cell(row=2, column=col).value
        if isinstance(v, datetime) and v.year >= 2000 and v.month == month:
            date_cols.append((col, v))

    rows_def = [
        (4, '午前', '会議室', ''),
        (5, '午前', '和室', '畳側'),
        (6, '午前', '洋室', '椅子側'),
        (7, '午前', '図書室', ''),
        (9, '午後', '会議室', ''),
        (10, '午後', '和室', '畳側'),
        (11, '午後', '洋室', '椅子側'),
        (12, '午後', '図書室', ''),
        (13, '夜間', '', ''),
    ]

    data = []
    for row_num, slot, room, sub_label in rows_def:
        cells = []
        for col, dt in date_cols:
            v = ws.cell(row=row_num, column=col).value
            s = str(v).strip() if v else ''
            s = s.replace('\u3000', '')
            cells.append(s)
        data.append({'slot': slot, 'room': room, 'sub': sub_label, 'cells': cells})

    return {
        'year': year, 'month': month, 'memo': str(memo),
        'dates': [dt for _, dt in date_cols], 'rows': data,
    }


def get_dow_name(dt):
    return DOW_NAMES[dt.weekday()]

def is_sunday(dt):
    return dt.weekday() == 6

def is_saturday(dt):
    return dt.weekday() == 5

def esc(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


def generate_html(month_data):
    year = month_data['year']
    month = month_data['month']
    memo = month_data['memo']
    dates = month_data['dates']
    rows = month_data['rows']
    num_days = len(dates)
    part_dates = dates
    total_cols = 3 + num_days + 1

    html = []
    html.append('<!DOCTYPE html>')
    html.append('<html lang="ja">')
    html.append('<head>')
    html.append('<meta charset="UTF-8">')
    html.append('<meta name="viewport" content="width=device-width, initial-scale=1.0">')
    html.append(f'<title>自治会館使用日程表 {year}年{month}月</title>')
    html.append('<style>' + get_css() + '</style>')
    html.append('</head><body>')

    html.append(f'<div class="table-section"><table>')

    # colgroup
    html.append('<colgroup>')
    html.append('<col style="width:28px"><col style="width:28px"><col style="width:24px">')
    for dt in part_dates:
        html.append('<col style="width:22px">')
        if dt.day == HALF:
            html.append('<col class="spacer-col" style="width:6px">')
    html.append('</colgroup>')

    # thead
    html.append('<thead>')
    html.append(f'<tr><th class="title-row" colspan="{total_cols}">自治会館使用日程表　　{year} 年 {month} 月</th></tr>')
    if memo:
        memo_text = memo if memo.startswith('備考') else '備考：' + memo
        html.append(f'<tr><td class="memo-row" colspan="{total_cols}">{esc(memo_text)}</td></tr>')

    # 日付行
    html.append('<tr class="date-row"><th class="corner" colspan="3">日</th>')
    for dt in part_dates:
        cls = 'sun' if is_sunday(dt) else ('sat' if is_saturday(dt) else '')
        html.append(f'<th class="day-cell {cls}">{dt.day}</th>')
        if dt.day == HALF:
            html.append('<th class="spacer"></th>')
    html.append('</tr>')

    # 曜日行
    html.append('<tr class="dow-row"><th class="corner" colspan="3">（曜）</th>')
    for dt in part_dates:
        dow = get_dow_name(dt)
        cls = 'sun' if is_sunday(dt) else ('sat' if is_saturday(dt) else '')
        html.append(f'<th class="dow-cell {cls}">{dow}</th>')
        if dt.day == HALF:
            html.append('<th class="spacer"></th>')
    html.append('</tr></thead>')

    # tbody
    html.append('<tbody>')
    prev_slot = ''
    for row in rows:
        slot, room, sub = row['slot'], row['room'], row['sub']
        bg = COLORS.get(room, COLORS.get('夜間', '#FFFFFF')) if room else COLORS.get(slot, '#FFFFFF')

        if slot != prev_slot and prev_slot and slot:
            if prev_slot == '午前' and slot == '午後':
                html.append(f'<tr class="separator"><td colspan="{total_cols}"></td></tr>')

        html.append('<tr>')
        if slot != prev_slot and slot:
            count = sum(1 for r in rows if r['slot'] == slot)
            html.append(f'<th class="slot-label" rowspan="{count}">{slot}</th>')
            prev_slot = slot

        if room:
            lbl_bg = LABEL_COLORS.get(room, bg)
            html.append(f'<th class="room-label" style="background:{lbl_bg}">{room}</th>')
            html.append(f'<th class="room-sub" style="background:{lbl_bg}">{sub}</th>')
        elif slot == '夜間':
            html.append(f'<th class="room-label" style="background:{bg}" colspan="2"></th>')
        else:
            html.append('<th class="room-label" colspan="2"></th>')

        for j, dt in enumerate(part_dates):
            val = row['cells'][j] if j < len(row['cells']) else ''
            cell_bg = bg
            pass
            content = f'<span class="closed-mark">×</span>' if val == '×' else esc(val)
            html.append(f'<td style="background:{cell_bg}">{content}</td>')
            if dt.day == HALF:
                html.append('<td class="spacer"></td>')
        html.append('</tr>')

    # フッター
    html.append(f'<tr><td class="footer-row" colspan="{total_cols}">')
    html.append(f'更新日: {datetime.now().strftime("%Y年%m月%d日")}　　（自治会館 Tel.045-784-4447）事務局員勤務、月〜土、9時〜12時')
    html.append('</td></tr>')

    html.append('</tbody></table></div></body></html>')
    return '\n'.join(html)


def get_css():
    return """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: "ＭＳ ゴシック", "MS Gothic", "Hiragino Kaku Gothic ProN", monospace; font-size: 13px; color: #000; background: #fff; padding: 16px; }
.title-row { background: #fff; font-size: 22px; font-weight: bold; padding: 8px; text-align: left; writing-mode: horizontal-tb; height: auto; border: none; }
.memo-row { background: #fff; font-size: 11px; color: #CC0000; padding: 4px 8px; text-align: left; writing-mode: horizontal-tb; height: auto; white-space: normal; border: none; }
.table-section { margin-bottom: 20px; overflow-x: auto; }
table { border-collapse: collapse; table-layout: fixed; }
th, td { border: 1px solid #999; text-align: center; vertical-align: middle; font-size: 11px; line-height: 1.3; }
.corner { background: #FFFFFF; padding: 2px 4px; white-space: nowrap; writing-mode: horizontal-tb; font-size: 11px; height: 20px; }
.slot-label { background: #FFFFFF; font-weight: bold; padding: 6px 4px; writing-mode: vertical-rl; text-orientation: upright; letter-spacing: 4px; font-size: 14px; white-space: nowrap; }
.room-label { font-weight: bold; padding: 2px 2px; white-space: nowrap; font-size: 12px; writing-mode: vertical-rl; text-orientation: upright; letter-spacing: 2px; min-width: 24px; }
.room-sub { font-size: 12px; padding: 2px 2px; white-space: nowrap; color: #555; writing-mode: vertical-rl; text-orientation: upright; min-width: 20px; }
.day-cell, .dow-cell { background: #FFFFFF; font-weight: bold; padding: 2px 2px; writing-mode: horizontal-tb; min-width: 28px; width: 28px; }
.day-cell.sun, .dow-cell.sun { color: #FF0000; background: #FFFFFF; }
.day-cell.sat, .dow-cell.sat { color: #0000FF; background: #FFFFFF; }
td { min-width: 24px; font-size: 11px; font-weight: bold; writing-mode: vertical-rl; text-orientation: mixed; white-space: normal; padding: 3px 1px; letter-spacing: 0.5px; max-height: 70px; height: 70px; overflow: hidden; text-overflow: ellipsis; }
.closed-mark { color: #000; font-weight: 900; font-size: 16px; }
.spacer { border-top: none !important; border-bottom: none !important; border-left: none !important; border-right: none !important; background: #fff !important; width: 6px !important; min-width: 6px !important; max-width: 6px !important; padding: 0 !important; writing-mode: horizontal-tb; height: auto; }
.spacer-col { width: 6px; }
.separator td { border: none !important; height: 6px; padding: 0; background: #fff; writing-mode: horizontal-tb; }
.footer-row { background: #fff; font-size: 11px; color: #666; padding: 6px 8px; text-align: center; writing-mode: horizontal-tb; height: auto; white-space: nowrap; border: none; }
@media print { body { padding: 0; } .table-section { overflow: visible; } }
"""


# =============================================
# GitHub Pages へ push
# =============================================

def github_push(config, html, filename):
    """GitHub REST API でファイルをpush"""
    token = config['github_token']
    owner = config['github_owner']
    repo = config['github_repo']
    branch = config.get('github_branch', 'main')
    path_prefix = config.get('github_path_prefix', '')
    path = path_prefix + filename

    url = f'https://api.github.com/repos/{owner}/{repo}/contents/{path}'

    # 既存ファイルのsha取得
    sha = None
    req = urllib.request.Request(
        f'{url}?ref={branch}',
        headers={
            'Authorization': f'token {token}',
            'Accept': 'application/vnd.github.v3+json',
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            sha = json.loads(resp.read().decode())['sha']
    except urllib.error.HTTPError:
        pass

    # push
    payload = {
        'message': f'{filename} を更新 ({datetime.now().strftime("%Y/%m/%d %H:%M")})',
        'content': base64.b64encode(html.encode('utf-8')).decode(),
        'branch': branch,
    }
    if sha:
        payload['sha'] = sha

    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode(),
        headers={
            'Authorization': f'token {token}',
            'Accept': 'application/vnd.github.v3+json',
            'Content-Type': 'application/json',
        },
        method='PUT',
    )

    try:
        with urllib.request.urlopen(req) as resp:
            code = resp.getcode()
            log.info(f'GitHub push 成功: {path} ({code})')
            return True
    except urllib.error.HTTPError as e:
        log.error(f'GitHub push 失敗: {path} ({e.code}) {e.read().decode()}')
        return False


# =============================================
# メイン
# =============================================

def main():
    force = '--force' in sys.argv

    if not force and not has_changed():
        log.debug('変更なし、スキップ')
        return

    log.info('Excel変更を検出、更新開始')

    config = load_config()
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)

    # 当月・翌月を特定
    now = datetime.now()
    target_months = [(now.year, now.month)]
    # 翌月
    if now.month == 12:
        target_months.append((now.year + 1, 1))
    else:
        target_months.append((now.year, now.month + 1))

    # シート一覧から対象を探す
    sheets = find_month_sheets(wb)

    for year, month in target_months:
        matching = [(y, m, name) for y, m, name in sheets if y == year and m == month]
        if not matching:
            log.warning(f'{year}年{month}月のシートが見つかりません')
            continue

        _, _, sheet_name = matching[0]
        month_data = read_month_data(wb, sheet_name)
        html = generate_html(month_data)
        filename = f'{year}-{month:02d}.html'

        github_push(config, html, filename)
        log.info(f'{year}年{month}月 → {filename} 生成完了')

    save_state()
    log.info('更新完了')


if __name__ == '__main__':
    main()
